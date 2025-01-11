import openpyxl
import random
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Step 1: Read teacher data from a file
teacher_data = {}
with open("teachers.txt", "r") as file:
    for line in file:
        name, subject = line.strip("()\n").split("(")
        teacher_data[name.strip()] = subject.strip()

# Get unique subjects and teacher names
subjects = list(set(teacher_data.values()))
teachers = list(teacher_data.keys())

# Step 2: Ask the user to input absent teachers
absent_teachers = input("Enter the names of absent teachers, separated by commas: ").split(",")
absent_teachers = [name.strip() for name in absent_teachers]

# Remove absent teachers from the list
teachers = [teacher for teacher in teachers if teacher not in absent_teachers]

# Step 3: Generate light colors for subjects
def generate_light_color():
    r = random.randint(200, 255)
    g = random.randint(200, 255)
    b = random.randint(200, 255)
    return f"{r:02X}{g:02X}{b:02X}"

subject_colors = {subject: generate_light_color() for subject in subjects}
teacher_colors = {teacher: subject_colors[teacher_data[teacher]] for teacher in teachers}

# Step 4: Define classes, sections, and periods
classes = [
    "9A", "9B", "9C", "9D",
    "10A", "10B", "10C", "10D",
    "11A", "11B",
    "12A", "12B",
]
num_periods = 8
max_periods_per_teacher = 7  # Adjusted to 7 periods maximum

# Initialize timetable and tracking data
timetable = {class_section: ["" for _ in range(num_periods)] for class_section in classes}
teacher_period_count = {teacher: 0 for teacher in teachers}
class_subject_count = {
    class_section: {subject: 0 for subject in subjects} for class_section in classes
}
teacher_assigned_periods = {teacher: set() for teacher in teachers}  # Track periods assigned to avoid conflicts

# Step 5: Assign teachers ensuring all conditions are met
def assign_teacher(class_name, period, used_teachers):
    random.shuffle(teachers)  # Shuffle for random assignment
    for teacher in teachers:
        subject = teacher_data[teacher]

        # Check conditions:
        # 1. Teacher must not exceed max periods
        # 2. Teacher must not teach in consecutive periods in the same class
        # 3. Teacher must not be teaching in another class during the same period
        # 4. At least one period of each subject must be assigned in the day
        if (
            teacher_period_count[teacher] < max_periods_per_teacher
            and (period == 0 or timetable[class_name][period - 1] != teacher)
            and period not in teacher_assigned_periods[teacher]
            and class_subject_count[class_name][subject] < 1
            and teacher not in used_teachers
        ):
            return teacher
    return None

# Step 6: Assign teachers to each period of each class
for class_name in classes:
    used_teachers_in_period = set()
    for period in range(num_periods):
        teacher = assign_teacher(class_name, period, used_teachers_in_period)
        if teacher:
            subject = teacher_data[teacher]
            timetable[class_name][period] = teacher
            teacher_period_count[teacher] += 1
            class_subject_count[class_name][subject] += 1
            teacher_assigned_periods[teacher].add(period)
            used_teachers_in_period.add(teacher)

# Step 7: Fill remaining empty periods with available teachers
for class_name, class_timetable in timetable.items():
    for period in range(num_periods):
        if class_timetable[period] == "":
            for teacher in teachers:
                if teacher_period_count[teacher] < max_periods_per_teacher:
                    subject = teacher_data[teacher]
                    if (
                        period not in teacher_assigned_periods[teacher]
                        and (period == 0 or class_timetable[period - 1] != teacher)
                        and class_subject_count[class_name][subject] < 1
                    ):
                        class_timetable[period] = teacher
                        teacher_period_count[teacher] += 1
                        class_subject_count[class_name][subject] += 1
                        teacher_assigned_periods[teacher].add(period)
                        break

# Step 8: Create Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Timetable"

# Define styles
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center")
header_font = Font(bold=True)

# Write header row
sheet["A1"] = "Period/Class"
sheet["A1"].fill = header_fill
sheet["A1"].font = header_font
sheet["A1"].alignment = alignment
sheet["A1"].border = border

for i, class_section in enumerate(classes):
    col = get_column_letter(i + 2)
    sheet[f"{col}1"] = class_section
    sheet[f"{col}1"].fill = header_fill
    sheet[f"{col}1"].font = header_font
    sheet[f"{col}1"].alignment = alignment
    sheet[f"{col}1"].border = border

# Fill timetable data
for period in range(num_periods):
    sheet[f"A{period + 2}"] = f"Period {period + 1}"
    sheet[f"A{period + 2}"].alignment = alignment
    sheet[f"A{period + 2}"].border = border
    for class_num, class_name in enumerate(classes):
        teacher = timetable[class_name][period]
        col = get_column_letter(class_num + 2)
        cell = sheet[f"{col}{period + 2}"]
        cell.value = teacher
        if teacher:
            cell.fill = PatternFill(start_color=teacher_colors[teacher], fill_type="solid")
        cell.alignment = alignment
        cell.border = border

# Step 9: Save the file
workbook.save("error_free_timetable.xlsx")
print("Timetable saved as 'error_free_timetable.xlsx'!")